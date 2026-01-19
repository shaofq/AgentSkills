# -*- coding: utf-8 -*-
"""船员信息比对服务"""
import os
import json
import pandas as pd
from typing import Dict, Any, List, Optional
from pathlib import Path
from datetime import datetime
from difflib import SequenceMatcher
from openai import OpenAI

from api.services.passport_ocr_service import get_passport_ocr_service

# 历史记录存储路径
HISTORY_DIR = Path("/tmp/crew_compare_history")
HISTORY_DIR.mkdir(parents=True, exist_ok=True)


class CrewCompareService:
    """
    船员信息比对服务：
    1. 解析 Excel 船员名单
    2. 识别护照图片
    3. 自动匹配和比对
    4. 生成比对报告
    """
    
    # Excel 列名映射（支持中英文）
    COLUMN_MAPPING = {
        # 英文列名
        'name': 'name',
        'family name': 'name',
        'given name': 'given_name',
        'full name': 'name',
        'nationality': 'nationality',
        'date of birth': 'date_of_birth',
        'birth date': 'date_of_birth',
        'sex': 'sex',
        'gender': 'sex',
        'place of birth': 'place_of_birth',
        'passport no': 'passport_no',
        'passport number': 'passport_no',
        'document no': 'passport_no',
        # 中文列名
        '姓名': 'name',
        '姓': 'surname',
        '名': 'given_name',
        '国籍': 'nationality',
        '出生日期': 'date_of_birth',
        '性别': 'sex',
        '出生地点': 'place_of_birth',
        '出生地': 'place_of_birth',
        '证件号码': 'passport_no',
        '护照号': 'passport_no',
        '护照号码': 'passport_no',
        '职务': 'rank',
        '登船日期': 'embark_date',
        '登船口岸': 'embark_port',
    }
    
    # 默认比对字段配置
    DEFAULT_COMPARE_FIELDS = [
        {"excel_field": "name", "passport_field": "full_name", "label": "姓名", "enabled": True},
        {"excel_field": "passport_no", "passport_field": "passport_no", "label": "证件号码", "enabled": True},
        {"excel_field": "nationality", "passport_field": "nationality", "label": "国籍", "enabled": True},
        {"excel_field": "date_of_birth", "passport_field": "date_of_birth", "label": "出生日期", "enabled": True},
        {"excel_field": "sex", "passport_field": "sex", "label": "性别", "enabled": True},
        {"excel_field": "place_of_birth", "passport_field": "place_of_birth", "label": "出生地点", "enabled": False},
    ]
    
    def __init__(self):
        self.passport_service = get_passport_ocr_service()
        self.sessions: Dict[str, Dict] = {}  # session_id -> session_data
        self.history: List[Dict] = []  # 操作历史记录
        self._load_history()
        
        # 初始化LLM客户端用于智能比对
        self.llm_client = OpenAI(
            api_key=os.getenv("DASHSCOPE_API_KEY"),
            base_url="https://dashscope.aliyuncs.com/compatible-mode/v1"
        )
        self.llm_model = "qwen-plus"  # 使用qwen-plus进行语义比对
    
    def create_session(self) -> str:
        """创建新的比对会话"""
        import uuid
        session_id = str(uuid.uuid4())[:8]
        self.sessions[session_id] = {
            "id": session_id,
            "created_at": datetime.now().isoformat(),
            "excel_data": None,
            "excel_file": None,
            "excel_columns": [],  # Excel 原始列名
            "passports": {},  # passport_no -> passport_data
            "compare_results": [],
            "compare_fields": self.DEFAULT_COMPARE_FIELDS.copy(),  # 可配置的比对字段
            "status": "created"
        }
        self._add_history(session_id, "create_session", "创建比对会话")
        print(f"[CrewCompare] 创建会话: {session_id}")
        return session_id
    
    def get_session(self, session_id: str) -> Optional[Dict]:
        """获取会话数据"""
        return self.sessions.get(session_id)
    
    # ============= 历史记录相关方法 =============
    
    def _load_history(self):
        """加载历史记录"""
        history_file = HISTORY_DIR / "history.json"
        if history_file.exists():
            try:
                with open(history_file, 'r', encoding='utf-8') as f:
                    self.history = json.load(f)
                print(f"[CrewCompare] 加载历史记录: {len(self.history)} 条")
            except Exception as e:
                print(f"[CrewCompare] 加载历史记录失败: {e}")
                self.history = []
        else:
            self.history = []
    
    def _save_history(self):
        """保存历史记录"""
        history_file = HISTORY_DIR / "history.json"
        try:
            with open(history_file, 'w', encoding='utf-8') as f:
                json.dump(self.history[-100:], f, ensure_ascii=False, indent=2)  # 只保留最近100条
        except Exception as e:
            print(f"[CrewCompare] 保存历史记录失败: {e}")
    
    def _add_history(self, session_id: str, action: str, detail: str, extra: Dict = None):
        """添加历史记录"""
        record = {
            "id": len(self.history) + 1,
            "session_id": session_id,
            "action": action,
            "detail": detail,
            "timestamp": datetime.now().isoformat(),
            "extra": extra or {}
        }
        self.history.append(record)
        self._save_history()
    
    def _save_session_snapshot(self, session_id: str):
        """保存会话快照到磁盘（用于历史记录恢复）"""
        session = self.sessions.get(session_id)
        if not session:
            return
        
        snapshot_dir = HISTORY_DIR / "sessions"
        snapshot_dir.mkdir(parents=True, exist_ok=True)
        snapshot_file = snapshot_dir / f"{session_id}.json"
        
        # 保存会话摘要（不包含大数据）
        snapshot = {
            "id": session["id"],
            "created_at": session.get("created_at"),
            "status": session.get("status"),
            "excel_file": session.get("excel_file"),
            "excel_filename": Path(session.get("excel_file", "")).name if session.get("excel_file") else None,
            "crew_count": len(session.get("excel_data", [])),
            "passport_count": len(session.get("passports", {})),
            "passport_files": list(session.get("passports", {}).keys()),
            "compare_results_count": len(session.get("compare_results", [])),
            "stats": session.get("stats"),
            "report_file": session.get("report_file"),
        }
        
        try:
            with open(snapshot_file, 'w', encoding='utf-8') as f:
                json.dump(snapshot, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"[CrewCompare] 保存会话快照失败: {e}")
    
    def get_session_snapshot(self, session_id: str) -> Optional[Dict]:
        """获取会话快照"""
        # 先尝试从内存获取
        session = self.sessions.get(session_id)
        if session:
            return {
                "id": session["id"],
                "created_at": session.get("created_at"),
                "status": session.get("status"),
                "excel_file": session.get("excel_file"),
                "excel_filename": Path(session.get("excel_file", "")).name if session.get("excel_file") else None,
                "crew_count": len(session.get("excel_data", [])),
                "passport_count": len(session.get("passports", {})),
                "passport_files": list(session.get("passports", {}).keys()),
                "compare_results_count": len(session.get("compare_results", [])),
                "stats": session.get("stats"),
                "report_file": session.get("report_file"),
            }
        
        # 从磁盘加载
        snapshot_file = HISTORY_DIR / "sessions" / f"{session_id}.json"
        if snapshot_file.exists():
            try:
                with open(snapshot_file, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except Exception as e:
                print(f"[CrewCompare] 加载会话快照失败: {e}")
        
        return None
    
    def get_session_files(self, session_id: str) -> Dict[str, Any]:
        """获取会话的所有文件列表"""
        upload_dir = Path(f"/tmp/crew_compare_uploads/{session_id}")
        
        result = {
            "excel_files": [],
            "passport_files": [],
            "report_files": []
        }
        
        if not upload_dir.exists():
            return result
        
        # Excel文件
        for f in upload_dir.glob("*.xlsx"):
            result["excel_files"].append({
                "filename": f.name,
                "path": str(f),
                "size": f.stat().st_size,
                "modified": datetime.fromtimestamp(f.stat().st_mtime).isoformat()
            })
        for f in upload_dir.glob("*.xls"):
            result["excel_files"].append({
                "filename": f.name,
                "path": str(f),
                "size": f.stat().st_size,
                "modified": datetime.fromtimestamp(f.stat().st_mtime).isoformat()
            })
        
        # 护照文件（图片和PDF转换的图片）
        passport_dir = upload_dir / "passports"
        if passport_dir.exists():
            for ext in ["*.jpg", "*.jpeg", "*.png", "*.bmp", "*.webp"]:
                for f in passport_dir.glob(ext):
                    result["passport_files"].append({
                        "filename": f.name,
                        "path": str(f),
                        "size": f.stat().st_size,
                        "modified": datetime.fromtimestamp(f.stat().st_mtime).isoformat()
                    })
            # PDF源文件
            for f in passport_dir.glob("*.pdf"):
                result["passport_files"].append({
                    "filename": f.name,
                    "path": str(f),
                    "size": f.stat().st_size,
                    "type": "pdf",
                    "modified": datetime.fromtimestamp(f.stat().st_mtime).isoformat()
                })
        
        # 报告文件
        for f in upload_dir.glob("*report*.xlsx"):
            result["report_files"].append({
                "filename": f.name,
                "path": str(f),
                "size": f.stat().st_size,
                "modified": datetime.fromtimestamp(f.stat().st_mtime).isoformat()
            })
        
        return result
    
    def get_history(self, limit: int = 50) -> List[Dict]:
        """获取历史记录"""
        return self.history[-limit:][::-1]  # 返回最近的记录，倒序
    
    def get_session_history(self, session_id: str) -> List[Dict]:
        """获取指定会话的历史记录"""
        return [h for h in self.history if h.get("session_id") == session_id]
    
    # ============= 比对字段配置方法 =============
    
    def get_compare_fields(self, session_id: str) -> List[Dict]:
        """获取会话的比对字段配置"""
        session = self.sessions.get(session_id)
        if not session:
            return self.DEFAULT_COMPARE_FIELDS.copy()
        return session.get("compare_fields", self.DEFAULT_COMPARE_FIELDS.copy())
    
    def update_compare_fields(self, session_id: str, fields: List[Dict]) -> Dict[str, Any]:
        """更新比对字段配置"""
        session = self.sessions.get(session_id)
        if not session:
            return {"error": "会话不存在"}
        
        session["compare_fields"] = fields
        enabled_fields = [f["label"] for f in fields if f.get("enabled")]
        self._add_history(session_id, "update_compare_fields", f"更新比对字段: {', '.join(enabled_fields)}")
        
        return {"success": True, "fields": fields}
    
    def get_available_excel_columns(self, session_id: str) -> List[str]:
        """获取Excel中可用的列名（用于字段映射配置）"""
        session = self.sessions.get(session_id)
        if not session or not session.get("excel_columns"):
            return []
        return session["excel_columns"]
    
    # ============= Excel列映射配置方法 =============
    
    def get_column_mapping(self, session_id: str) -> Dict[str, Any]:
        """获取当前列映射配置"""
        session = self.sessions.get(session_id)
        if not session:
            return {"error": "会话不存在"}
        
        # 返回原始列名、标准化列名和当前映射
        return {
            "success": True,
            "original_columns": session.get("original_columns", []),
            "column_mapping": session.get("column_mapping", {}),
            "available_fields": list(self.COLUMN_MAPPING.values())
        }
    
    def update_column_mapping(self, session_id: str, mapping: Dict[str, str]) -> Dict[str, Any]:
        """更新列映射配置并重新解析数据"""
        session = self.sessions.get(session_id)
        if not session:
            return {"error": "会话不存在"}
        
        session["column_mapping"] = mapping
        
        # 如果已有Excel数据，重新应用映射
        if session.get("raw_excel_data"):
            crew_list = self._apply_column_mapping(session["raw_excel_data"], mapping)
            session["excel_data"] = crew_list
            self._add_history(session_id, "update_column_mapping", f"更新列映射配置")
            return {"success": True, "crew_list": crew_list, "count": len(crew_list)}
        
        return {"success": True, "message": "映射已保存，将在下次解析时生效"}
    
    def _apply_column_mapping(self, raw_data: List[Dict], mapping: Dict[str, str]) -> List[Dict]:
        """应用列映射到原始数据"""
        crew_list = []
        for i, record in enumerate(raw_data):
            crew = {"index": i + 1}
            for orig_col, value in record.items():
                # 使用自定义映射或默认映射
                target_field = mapping.get(orig_col) or self._normalize_column(orig_col)
                if target_field and value is not None:
                    crew[target_field] = str(value).strip() if pd.notna(value) else ""
            
            non_empty = sum(1 for k, v in crew.items() if v and k != 'index')
            if non_empty > 0:
                crew_list.append(crew)
        return crew_list
    
    # ============= 护照识别结果编辑方法 =============
    
    def update_passport_result(self, session_id: str, passport_no: str, updates: Dict[str, Any]) -> Dict[str, Any]:
        """手动编辑护照识别结果"""
        session = self.sessions.get(session_id)
        if not session:
            return {"error": "会话不存在"}
        
        passports = session.get("passports", {})
        
        # 查找护照（可能key是原始护照号或文件名）
        target_passport = None
        target_key = None
        for key, passport in passports.items():
            if key == passport_no or passport.get("passport_no") == passport_no:
                target_passport = passport
                target_key = key
                break
        
        if not target_passport:
            return {"error": f"未找到护照: {passport_no}"}
        
        # 更新字段
        old_values = {}
        for field, new_value in updates.items():
            if field in target_passport:
                old_values[field] = target_passport[field]
            target_passport[field] = new_value
        
        # 如果护照号被修改，更新字典key
        if "passport_no" in updates and updates["passport_no"] != target_key:
            del passports[target_key]
            passports[updates["passport_no"]] = target_passport
        
        target_passport["manually_edited"] = True
        target_passport["edit_time"] = datetime.now().isoformat()
        
        self._add_history(session_id, "edit_passport", f"编辑护照: {passport_no}", {"updates": updates})
        
        return {"success": True, "passport": target_passport}
    
    def recognize_single_passport(self, session_id: str, filename: str) -> Dict[str, Any]:
        """重新识别单张护照"""
        session = self.sessions.get(session_id)
        if not session:
            return {"error": "会话不存在"}
        
        # 查找护照文件路径
        passport_dir = Path(f"/tmp/crew_compare_uploads/{session_id}/passports")
        file_path = passport_dir / filename
        
        if not file_path.exists():
            return {"error": f"文件不存在: {filename}"}
        
        try:
            # 调用OCR识别
            result = self.passport_service.recognize(str(file_path))
            
            if result.get("error"):
                return {"error": result["error"]}
            
            # 更新护照数据
            passport_no = result.get("passport_no", filename)
            session["passports"][passport_no] = result
            result["filename"] = filename
            result["recognized_at"] = datetime.now().isoformat()
            
            self._add_history(session_id, "re_recognize_passport", f"重新识别护照: {filename}")
            
            return {"success": True, "result": result}
            
        except Exception as e:
            return {"error": str(e)}
    
    def parse_excel(self, session_id: str, file_path: str) -> Dict[str, Any]:
        """
        解析 Excel 船员名单。
        
        Args:
            session_id: 会话 ID
            file_path: Excel 文件路径
            
        Returns:
            解析结果
        """
        session = self.sessions.get(session_id)
        if not session:
            return {"error": "会话不存在"}
        
        try:
            # 读取 Excel，尝试多种方式
            print(f"[CrewCompare] 读取文件: {file_path}")
            
            # 先尝试读取所有 sheet 名称
            xl = pd.ExcelFile(file_path)
            print(f"[CrewCompare] Sheet 列表: {xl.sheet_names}")
            
            # 读取第一个 sheet
            df = pd.read_excel(file_path, sheet_name=0)
            
            # 打印原始数据信息
            print(f"[CrewCompare] 数据形状: {df.shape}")
            print(f"[CrewCompare] 原始列名: {list(df.columns)}")
            print(f"[CrewCompare] 前3行数据: {df.head(3).to_dict('records')}")
            
            # 移除空行
            df = df.dropna(how='all')
            
            # 保存原始列名用于列映射配置
            original_columns = [str(col) for col in df.columns]
            session["original_columns"] = original_columns
            
            # 标准化列名（保留原始列名作为备用）
            column_map = {}
            for col in df.columns:
                normalized = self._normalize_column(col)
                column_map[str(col)] = normalized
            
            session["column_mapping"] = column_map  # 保存列映射
            
            # 保存原始数据用于后续重新映射
            raw_records = df.to_dict('records')
            session["raw_excel_data"] = raw_records
            
            # 应用列映射
            df.columns = [column_map.get(str(col), str(col)) for col in df.columns]
            print(f"[CrewCompare] 标准化列名: {list(df.columns)}")
            
            # 转换为记录列表
            records = df.to_dict('records')
            
            # 处理每条记录 - 放宽条件，只要有任何数据就保留
            crew_list = []
            for i, record in enumerate(records):
                crew = self._process_crew_record(record, i + 1)
                # 只要记录有超过1个非空字段就保留
                non_empty_fields = sum(1 for k, v in crew.items() if v and k != 'index')
                if non_empty_fields > 0:
                    crew_list.append(crew)
            
            session["excel_data"] = crew_list
            session["excel_file"] = file_path
            session["excel_columns"] = list(df.columns)  # 保存列名供配置使用
            session["status"] = "excel_loaded"
            
            # 记录历史
            self._add_history(session_id, "upload_excel", f"上传Excel: {Path(file_path).name}, {len(crew_list)} 条记录")
            self._save_session_snapshot(session_id)
            
            print(f"[CrewCompare] 解析 Excel 完成: {len(crew_list)} 条记录")
            
            return {
                "success": True,
                "count": len(crew_list),
                "columns": list(df.columns),
                "crew_list": crew_list,
                "compare_fields": session.get("compare_fields", self.DEFAULT_COMPARE_FIELDS)
            }
            
        except Exception as e:
            print(f"[CrewCompare] 解析 Excel 失败: {e}")
            return {"error": str(e)}
    
    def _normalize_column(self, col: str) -> str:
        """标准化列名 - 处理复杂的中英文混合列名"""
        col_str = str(col).strip()
        
        # 尝试直接匹配
        col_lower = col_str.lower()
        if col_lower in self.COLUMN_MAPPING:
            return self.COLUMN_MAPPING[col_lower]
        
        # 移除特殊字符后匹配
        cleaned = col_str.replace('*', '').replace('\n', ' ').strip().lower()
        if cleaned in self.COLUMN_MAPPING:
            return self.COLUMN_MAPPING[cleaned]
        
        # 按关键词匹配
        col_check = col_str.lower()
        
        # 姓名
        if '姓名' in col_check or 'family name' in col_check or 'given name' in col_check:
            return 'name'
        # 证件号码/护照号
        if '证件号码' in col_check or 'identity document' in col_check or 'passport' in col_check:
            return 'passport_no'
        # 国籍
        if '国籍' in col_check or 'nationality' in col_check:
            return 'nationality'
        # 出生日期
        if '出生日期' in col_check or 'date of birth' in col_check or 'birth' in col_check:
            return 'date_of_birth'
        # 性别
        if '性别' in col_check or 'm/f' in col_check or 'gender' in col_check or 'sex' in col_check:
            return 'sex'
        # 出生地点
        if '出生地' in col_check or 'place of birth' in col_check:
            return 'place_of_birth'
        # 职务
        if '职务' in col_check or 'rank' in col_check or 'rating' in col_check:
            return 'rank'
        # 证件类型
        if '证件类型' in col_check:
            return 'document_type'
        # 登船口岸
        if '登船口岸' in col_check:
            return 'embark_port'
        # 登船日期
        if '登船日期' in col_check:
            return 'embark_date'
        
        return cleaned
    
    def _process_crew_record(self, record: Dict, index: int) -> Dict:
        """处理单条船员记录"""
        crew = {"index": index}
        
        for key, value in record.items():
            if pd.isna(value):
                continue
            
            # 标准化字段名
            field = self._normalize_column(key)
            
            # 处理日期字段
            if 'date' in field and value:
                value = self._format_date(value)
            
            # 处理性别
            if field == 'sex' and value:
                value = self._normalize_sex(str(value))
            
            # 处理姓名（大写）
            if field in ['name', 'surname', 'given_name'] and value:
                value = str(value).upper().strip()
            
            crew[field] = value
        
        # 合并姓名
        if 'surname' in crew and 'given_name' in crew:
            crew['name'] = f"{crew.get('surname', '')} {crew.get('given_name', '')}".strip()
        
        return crew
    
    def _format_date(self, value) -> str:
        """格式化日期，统一转为 YYYY-MM-DD 格式"""
        if not value:
            return ""
        
        if isinstance(value, datetime):
            return value.strftime('%Y-%m-%d')
        
        # 处理整数类型的日期（如 19751022）
        if isinstance(value, (int, float)):
            value = str(int(value))
        
        value = str(value).strip()
        
        # 尝试解析常见格式
        formats = [
            '%Y-%m-%d',      # 1975-10-22
            '%Y/%m/%d',      # 1975/10/22
            '%d/%m/%Y',      # 22/10/1975
            '%Y%m%d',        # 19751022
            '%d %b %Y',      # 22 OCT 1975
            '%d %B %Y',      # 22 October 1975
        ]
        
        for fmt in formats:
            try:
                dt = datetime.strptime(value.upper(), fmt)
                return dt.strftime('%Y-%m-%d')
            except:
                continue
        
        # 尝试处理 8 位数字格式
        if len(value) == 8 and value.isdigit():
            try:
                dt = datetime.strptime(value, '%Y%m%d')
                return dt.strftime('%Y-%m-%d')
            except:
                pass
        
        return value
    
    def _normalize_sex(self, value: str) -> str:
        """标准化性别"""
        value = value.upper().strip()
        if value in ['M', '男', 'MALE']:
            return 'M'
        if value in ['F', '女', 'FEMALE']:
            return 'F'
        return value
    
    async def add_passport(self, session_id: str, image_path: str) -> Dict[str, Any]:
        """
        添加并识别护照图片。
        
        Args:
            session_id: 会话 ID
            image_path: 护照图片路径
            
        Returns:
            识别结果
        """
        session = self.sessions.get(session_id)
        if not session:
            return {"error": "会话不存在"}
        
        # 识别护照
        result = await self.passport_service.recognize_passport(image_path)
        
        if "error" not in result:
            # 以护照号为 key 存储
            passport_no = result.get("passport_no", "")
            if passport_no:
                session["passports"][passport_no] = result
            else:
                # 没有护照号则用文件名
                session["passports"][Path(image_path).stem] = result
            
            session["status"] = "passports_added"
        
        return result
    
    def compare(self, session_id: str) -> Dict[str, Any]:
        """
        执行比对。
        
        Args:
            session_id: 会话 ID
            
        Returns:
            比对结果
        """
        session = self.sessions.get(session_id)
        if not session:
            return {"error": "会话不存在"}
        
        excel_data = session.get("excel_data", [])
        passports = session.get("passports", {})
        
        if not excel_data:
            return {"error": "未上传 Excel 名单"}
        
        if not passports:
            return {"error": "未上传护照图片"}
        
        results = []
        matched_passports = set()
        
        # 遍历 Excel 记录进行匹配
        # 获取比对字段配置
        compare_fields = session.get("compare_fields", self.DEFAULT_COMPARE_FIELDS)
        
        for crew in excel_data:
            match_result = self._match_and_compare(crew, passports, matched_passports, compare_fields)
            results.append(match_result)
        
        # 统计
        stats = {
            "total": len(excel_data),
            "matched": sum(1 for r in results if r["match_status"] == "matched"),
            "mismatched": sum(1 for r in results if r["match_status"] == "mismatched"),
            "not_found": sum(1 for r in results if r["match_status"] == "not_found"),
            "passport_count": len(passports),
            "unmatched_passports": len(passports) - len(matched_passports)
        }
        
        session["compare_results"] = results
        session["stats"] = stats
        session["status"] = "compared"
        
        self._add_history(session_id, "compare", f"比对完成: 匹配{stats['matched']}, 差异{stats['mismatched']}, 未找到{stats['not_found']}")
        self._save_session_snapshot(session_id)
        
        print(f"[CrewCompare] 比对完成: 匹配 {stats['matched']}, 差异 {stats['mismatched']}, 未找到 {stats['not_found']}")
        
        return {
            "success": True,
            "results": results,
            "stats": stats
        }
    
    def _match_and_compare(
        self, 
        crew: Dict, 
        passports: Dict[str, Dict],
        matched_passports: set,
        compare_fields: List[Dict] = None
    ) -> Dict[str, Any]:
        """匹配并比对单条记录"""
        result = {
            "crew": crew,
            "passport": None,
            "match_status": "not_found",
            "differences": [],
            "match_score": 0
        }
        
        crew_passport_no = str(crew.get("passport_no", "")).strip().upper()
        crew_name = str(crew.get("name", "")).strip().upper()
        
        # 优先按护照号匹配
        best_match = None
        best_score = 0
        
        for passport_no, passport in passports.items():
            if passport_no in matched_passports:
                continue
            
            passport_no_ocr = str(passport.get("passport_no", "")).strip().upper()
            passport_name = str(passport.get("full_name", "")).strip().upper()
            
            score = 0
            
            # 护照号完全匹配
            if crew_passport_no and passport_no_ocr:
                if crew_passport_no == passport_no_ocr:
                    score += 100
                elif self._similar(crew_passport_no, passport_no_ocr) > 0.8:
                    score += 80
            
            # 姓名匹配
            if crew_name and passport_name:
                name_sim = self._similar(crew_name, passport_name)
                score += name_sim * 50
            
            if score > best_score:
                best_score = score
                best_match = passport
        
        # 设置匹配结果
        if best_match and best_score >= 50:
            result["passport"] = best_match
            result["match_score"] = best_score
            matched_passports.add(best_match.get("passport_no", ""))
            
            # 比对字段差异（使用会话配置的字段）
            differences = self._compare_fields(crew, best_match, compare_fields)
            result["differences"] = differences
            
            if differences:
                result["match_status"] = "mismatched"
            else:
                result["match_status"] = "matched"
        
        return result
    
    def _similar(self, a: str, b: str) -> float:
        """计算字符串相似度"""
        return SequenceMatcher(None, a, b).ratio()
    
    def _llm_semantic_compare(self, fields_to_compare: List[Dict]) -> Dict[str, Dict]:
        """
        使用语言模型进行语义比对。
        
        Args:
            fields_to_compare: 待比对的字段列表，每项包含 field_name, excel_value, passport_value
            
        Returns:
            比对结果字典，key为field_name，value包含 is_match, reason
        """
        if not fields_to_compare:
            return {}
        
        # 构建比对prompt
        compare_items = []
        for item in fields_to_compare:
            compare_items.append(f"- {item['field_name']}: Excel值=\"{item['excel_value']}\" vs 护照值=\"{item['passport_value']}\"")
        
        prompt = f"""你是一个数据比对专家。请判断以下Excel数据和护照数据是否语义上一致。

需要考虑以下情况：
1. 语言差异：如"中国"与"CHINA"、"菲律宾"与"PHILIPPINES"是一致的
2. 缩写差异：如"男"与"M"、"女"与"F"是一致的
3. 姓名顺序：如"ZHANG SAN"与"SAN ZHANG"可能是同一个人
4. 日期格式：如"1990-01-15"与"15 JAN 1990"是一致的
5. 大小写差异：忽略大小写差异

待比对字段：
{chr(10).join(compare_items)}

请以JSON格式返回每个字段的比对结果，格式如下：
{{
  "字段名": {{"is_match": true/false, "reason": "判断理由"}},
  ...
}}

只返回JSON，不要其他内容。"""

        try:
            response = self.llm_client.chat.completions.create(
                model=self.llm_model,
                messages=[{"role": "user", "content": prompt}],
                temperature=0.1,
                max_tokens=1000
            )
            
            result_text = response.choices[0].message.content.strip()
            print(f"[CrewCompare] LLM语义比对结果: {result_text}")
            # 提取JSON
            if "```json" in result_text:
                result_text = result_text.split("```json")[1].split("```")[0].strip()
            elif "```" in result_text:
                result_text = result_text.split("```")[1].split("```")[0].strip()
            
            return json.loads(result_text)
            
        except Exception as e:
            print(f"[CrewCompare] LLM语义比对失败: {e}")
            return {}
    
    def _compare_fields(self, crew: Dict, passport: Dict, compare_fields: List[Dict] = None) -> List[Dict]:
        """比对字段差异，使用LLM智能语义比对"""
        differences = []
        
        # 使用传入的字段配置，否则使用默认配置
        if compare_fields is None:
            compare_fields = self.DEFAULT_COMPARE_FIELDS
        
        # 收集需要比对的字段
        fields_to_compare = []
        field_info_map = {}  # 存储字段原始信息
        
        for field_config in compare_fields:
            if not field_config.get("enabled", True):
                continue
            
            excel_field = field_config.get("excel_field")
            passport_field = field_config.get("passport_field")
            label = field_config.get("label")
            
            if not excel_field or not passport_field:
                continue
                
            excel_val = str(crew.get(excel_field, "")).strip()
            passport_val = str(passport.get(passport_field, "")).strip()
            
            if not excel_val or not passport_val:
                continue
            
            # 完全匹配（忽略大小写）则跳过
            if excel_val.upper() == passport_val.upper():
                continue
            
            # 不完全匹配的字段，交给LLM进行语义比对
            fields_to_compare.append({
                "field_name": label,
                "excel_value": excel_val,
                "passport_value": passport_val
            })
            field_info_map[label] = {
                "excel_field": excel_field,
                "passport_field": passport_field,
                "excel_val": excel_val,
                "passport_val": passport_val
            }
        
        # 使用LLM进行语义比对
        print(f"[Compare] 需要LLM比对的字段数: {len(fields_to_compare)}, 字段: {[f['field_name'] for f in fields_to_compare]}")
        if fields_to_compare:
            llm_results = self._llm_semantic_compare(fields_to_compare)
            print(f"[Compare] LLM比对结果: {llm_results}")
            
            for field_name, info in field_info_map.items():
                llm_result = llm_results.get(field_name, {})
                is_match = llm_result.get("is_match", False)
                reason = llm_result.get("reason", "")
                
                if not is_match:
                    # 计算字符串相似度
                    similarity = self._similar(info["excel_val"].upper(), info["passport_val"].upper())
                    
                    differences.append({
                        "field": field_name,
                        "excel_value": info["excel_val"],
                        "passport_value": info["passport_val"],
                        "similarity": round(similarity, 2),
                        "severity": "high" if similarity < 0.5 else "medium" if similarity < 0.8 else "low",
                        "ai_reason": reason  # 添加AI判断理由
                    })
        
        return differences
    
    def export_report(self, session_id: str, output_path: str) -> Dict[str, Any]:
        """
        导出比对报告为 Excel。
        
        Args:
            session_id: 会话 ID
            output_path: 输出文件路径
            
        Returns:
            导出结果
        """
        session = self.sessions.get(session_id)
        if not session:
            return {"error": "会话不存在"}
        
        results = session.get("compare_results", [])
        if not results:
            return {"error": "未进行比对"}
        
        # 构建报告数据
        report_data = []
        for r in results:
            crew = r.get("crew", {})
            passport = r.get("passport", {})
            
            row = {
                "序号": crew.get("index", ""),
                "姓名(Excel)": crew.get("name", ""),
                "姓名(护照)": passport.get("full_name", "") if passport else "",
                "证件号(Excel)": crew.get("passport_no", ""),
                "证件号(护照)": passport.get("passport_no", "") if passport else "",
                "国籍(Excel)": crew.get("nationality", ""),
                "国籍(护照)": passport.get("nationality", "") if passport else "",
                "出生日期(Excel)": crew.get("date_of_birth", ""),
                "出生日期(护照)": passport.get("date_of_birth", "") if passport else "",
                "比对状态": {
                    "matched": "✓ 一致",
                    "mismatched": "⚠ 有差异",
                    "not_found": "✗ 未找到护照"
                }.get(r.get("match_status"), ""),
                "差异详情": "; ".join([
                    f"{d['field']}: {d['excel_value']} → {d['passport_value']}"
                    for d in r.get("differences", [])
                ]) if r.get("differences") else "",
                "护照有效期": passport.get("date_of_expiry", "") if passport else "",
            }
            report_data.append(row)
        
        # 创建 DataFrame 并导出
        df = pd.DataFrame(report_data)
        
        # 使用 openpyxl 创建美化的 Excel
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl.utils import get_column_letter
        
        wb = Workbook()
        ws = wb.active
        ws.title = "船员护照比对报告"
        
        # 定义样式
        header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        data_font = Font(name='微软雅黑', size=10)
        data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # 边框样式
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )
        
        # 状态颜色
        status_fills = {
            '✓ 一致': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
            '⚠ 有差异': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),
            '✗ 未找到护照': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
        }
        status_fonts = {
            '✓ 一致': Font(name='微软雅黑', size=10, color='006100'),
            '⚠ 有差异': Font(name='微软雅黑', size=10, color='9C5700'),
            '✗ 未找到护照': Font(name='微软雅黑', size=10, color='9C0006'),
        }
        
        # 写入数据
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.border = thin_border
                
                if r_idx == 1:
                    # 表头样式
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                else:
                    # 数据行样式
                    cell.font = data_font
                    cell.alignment = data_alignment
                    
                    # 比对状态列特殊着色
                    col_name = df.columns[c_idx - 1] if c_idx <= len(df.columns) else ""
                    if col_name == "比对状态" and value in status_fills:
                        cell.fill = status_fills[value]
                        cell.font = status_fonts[value]
                        cell.alignment = center_alignment
                    elif col_name == "序号":
                        cell.alignment = center_alignment
        
        # 自适应列宽
        def calculate_text_width(text):
            """计算文本显示宽度，中文字符按2倍计算"""
            if not text:
                return 0
            text = str(text)
            width = 0
            for char in text:
                if '\u4e00' <= char <= '\u9fff' or '\u3000' <= char <= '\u303f':
                    width += 2.2  # 中文字符
                elif '\uff00' <= char <= '\uffef':
                    width += 2.2  # 全角字符
                else:
                    width += 1.1  # 英文/数字
            return width
        
        for col_idx, column in enumerate(df.columns, 1):
            # 计算表头宽度
            max_width = calculate_text_width(column)
            
            # 计算数据列宽度
            for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        cell_width = calculate_text_width(cell.value)
                        max_width = max(max_width, cell_width)
            
            # 设置列宽，增加边距，限制范围
            adjusted_width = max(8, min(max_width + 3, 60))
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
        
        # 设置行高
        ws.row_dimensions[1].height = 32  # 表头行高
        for row_idx in range(2, ws.max_row + 1):
            ws.row_dimensions[row_idx].height = 24
        
        # 冻结首行
        ws.freeze_panes = 'A2'
        
        # 保存文件
        wb.save(output_path)
        
        # 保存报告路径到会话
        session["report_file"] = output_path
        
        self._add_history(session_id, "export_report", f"导出报告: {Path(output_path).name}")
        self._save_session_snapshot(session_id)
        
        print(f"[CrewCompare] 报告已导出: {output_path}")
        
        return {
            "success": True,
            "file_path": output_path,
            "count": len(report_data)
        }


# 单例实例
_crew_compare_service: Optional[CrewCompareService] = None


def get_crew_compare_service() -> CrewCompareService:
    """获取船员比对服务单例"""
    global _crew_compare_service
    if _crew_compare_service is None:
        _crew_compare_service = CrewCompareService()
    return _crew_compare_service
